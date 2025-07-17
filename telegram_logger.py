# .\ngrok.exe config add-authtoken 2ztqmMqXwQCnkCc3N6fHvzJVVN8_4Nm3obS8M8ZexsCLzR9rs
# .\ngrok.exe http 5000


from flask import Flask, request, render_template_string, redirect, url_for, session
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = '9f3e8c2b5d7a4f9cbb8e1d0a3f7c6e4520d93f4a1b6c7e8d9f1a2b3c4d5e6f7a'

USERNAME = 'Marnus'
PASSWORD = 'NEX@test149'  # Change this!

EXCEL_FILE = 'C:/Users/marnus.muller/OneDrive - neXgro (Pty) Ltd/Documents/BOT/telegram_bot/fuel_log.xlsx'

HTML_FORM = '''
<!DOCTYPE html>
<html>
<head>
    <title>Fuel Log Entry</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <style>
        body {
            font-size: 20px;             /* Larger base font size */
            font-family: Arial, sans-serif;
            padding: 30px;
            max-width: 700px;            /* Wider container */
            margin: auto;
        }
        label {
            display: block;
            margin-top: 25px;
            font-weight: bold;
        }
        input, textarea {
            width: 100%;
            padding: 20px 20px;          /* Bigger padding */
            font-size: 22px;             /* Bigger font */
            box-sizing: border-box;
            border-radius: 30px;
            border: 2px solid #ccc;
        }
        textarea {
            height: 100px;               /* Taller textarea */
        }
        button {
            margin-top: 30px;
            padding: 70px;
            font-size: 28px;             /* Bigger button text */
            width: 100%;
            background-color: #4CAF50;
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            transition: background-color 0.3s ease;
        }
        button:hover {
            background-color: #45a049;
        }
        .result {
            font-weight: bold;
            margin-top: 26px;
            font-size: 26px;             /* Larger pumped display */
        }
        img.logo {
            max-width: 200px;
            display: block;
            margin: 0 auto 20px auto;
        }
        .logout {
            position: absolute;
            top: 20px;
            right: 30px;
            font-size: 18px;
        }
        .logout a {
            text-decoration: none;
            color: red;
            font-weight: bold;
        }
        .logout a:hover {
            text-decoration: underline;
        }

        select {
            width: 100%;
            padding: 20px;
            font-size: 22px;
            border-radius: 12px;
            border: 2px solid #ccc;
            margin-top: 15px;
            appearance: none;
            background-color: #fff;
        }

        select option {
            font-size: 22px;
            padding: 10px;
        }
    </style>

    <script>
        function calculateEnd() {
            var start = parseFloat(document.getElementById('start').value) || 0;
            var pumped = parseFloat(document.getElementById('pumped_input').value) || 0;
            var end = start + pumped;
            document.getElementById('calculated_end').textContent = end.toFixed(2);
        }
    </script>
</head>
<body>
    <img class="logo" src="/static/logo.png" alt="Logo">
    <h2>⛽ Fuel Log Entry</h2>
    <form method="POST">
        <!-- Site Dropdown -->
        <label for="site">Select Site:</label>
        <select id="site" name="site" onchange="toggleVehicleField()">
            <option value="">--Select Site--</option>
            <option value="Holfontein">Holfontein</option>
            <option value="Plank">Plank</option>
        </select>

        <!-- Holfontein Vehicle Dropdown -->
        <div id="vehicle_dropdown" style="display:none;">
            <label for="vehicle_select">Vehicle:</label>
            <select id="vehicle_select" name="vehicle_select">
                <option value="">--Select Vehicle--</option>
                <option>Geni 1</option>
                <option>Geni 2</option>
                <option>Geni3 Hopper</option>
                <option>Landini 1</option>
                <option>Landini 2</option>
                <option>Landini 3</option>
                <option>Landini 4</option>
                <option>Landini 5</option>
                <option>Mahindra Bakkie</option>
                <option>MF DHS856FS</option>
                <option>MF DHS872FS</option>
                <option>MF DHS879FS</option>
                <option>MF DHS885FS</option>
            </select>
        </div>

        <!-- Plank Vehicle Text Input -->
        <div id="vehicle_input" style="display:none;">
            <label for="vehicle_text">Vehicle:</label>
            <input id="vehicle_text" name="vehicle_text" type="text">
        </div>

        <script>
        function toggleVehicleField() {
            var site = document.getElementById("site").value;
            if (site === "Holfontein") {
                document.getElementById("vehicle_dropdown").style.display = "block";
                document.getElementById("vehicle_input").style.display = "none";
            } else if (site === "Plank") {
                document.getElementById("vehicle_dropdown").style.display = "none";
                document.getElementById("vehicle_input").style.display = "block";
            } else {
                document.getElementById("vehicle_dropdown").style.display = "none";
                document.getElementById("vehicle_input").style.display = "none";
            }
        }
        </script>


        <label for="driver_name">Driver Name:</label>
        <input id="driver_name" name="driver_name" type="text" required>

        <label for="odometer">Vehicle Odometer:</label>
        <input id="odometer" name="odometer" type="number" step="0.1" required>


        <label for="start">Pump Start Reading:</label>
        <input id="start" name="start" type="number" step="0.01" required oninput="calculateEnd()">
        {% if error %}
        <div style="color:red; font-size:18px; font-weight:bold; margin-top:5px;">{{ error }}</div>
        {% endif %}


        <label for="pumped_input">Pumped (Litres):</label>
        <input id="pumped_input" name="pumped" type="number" step="0.1" required oninput="calculateEnd()">

        <div class="result">End Reading: <span id="calculated_end">0.00</span></div>

        <button type="submit">Log Fuel</button>
    </form>
    <div class="logout"><a href="/logout">Logout</a></div>

</body>
</html>
'''

LOGIN_FORM = '''
<!DOCTYPE html>
<html>
<head>
    <title>Login</title>
    <meta name="viewport" content="width=device-width, initial-scale=1">  <!-- ✅ Important for mobile scaling -->
    <style>
        body {
            font-size: 22px;
            font-family: Arial, sans-serif;
            padding: 20px;
            margin: 0;
            background-color: #f9f9f9;
        }
        .container {
            max-width: 500px;
            margin: auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 0 10px rgba(0,0,0,0.1);
        }
        label {
            display: block;
            margin-top: 20px;
        }
        input {
            width: 100%;
            padding: 15px;
            font-size: 22px;
            box-sizing: border-box;
            border: 1.5px solid #ccc;
            border-radius: 6px;
        }
        button {
            margin-top: 30px;
            padding: 18px;
            font-size: 24px;
            width: 100%;
            background-color: #007BFF;
            color: white;
            border: none;
            border-radius: 6px;
            cursor: pointer;
        }
        button:hover {
            background-color: #0069d9;
        }
        .error {
            color: red;
            margin-top: 20px;
            font-size: 20px;
        }
        img.logo {
            max-width: 200px;
            display: block;
            margin: 0 auto 20px auto;
        }
    </style>
</head>
<body>
    <div class="container">
        <img class="logo" src="/static/logo.png" alt="Logo">
        <h2>🔐 Login</h2>
        <form method="POST">
            <label for="username">Username:</label>
            <input id="username" name="username" required autofocus>

            <label for="password">Password:</label>
            <input id="password" name="password" type="password" required>
            <input type="checkbox" onclick="togglePassword()"> Show Password

            <script>
            function togglePassword() {
                var pass = document.getElementById("password");
                pass.type = (pass.type === "password") ? "text" : "password";
            }
            </script>


            <button type="submit">Login</button>
        </form>
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
    </div>
</body>
</html>
'''


def write_to_excel(timestamp, site, vehicle, driver_name, odometer, start, end, pumped):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Site', 'Vehicle', 'Driver Name', 'Odometer', 'Start Reading', 'End Reading', 'Pumped'])

    ws.append([timestamp, site, vehicle, driver_name, odometer, start, end, pumped])
    wb.save(EXCEL_FILE)




def get_last_readings():
    if not os.path.exists(EXCEL_FILE):
        return None

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    if ws.max_row < 2:
        return None

    last_row = list(ws.iter_rows(min_row=ws.max_row, values_only=True))[0]
    last_start = float(last_row[5])  
    last_end = float(last_row[6])    
    return last_start, last_end


@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        if request.form['username'] == USERNAME and request.form['password'] == PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('log_fuel'))
        else:
            error = "Invalid username or password"
    return render_template_string(LOGIN_FORM, error=error)

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/', methods=['GET', 'POST'])
def log_fuel():
    if not session.get('logged_in'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        # ✅ Capture all fields from form
        site = request.form.get('site')
        driver_name = request.form.get('driver_name')
        odometer = float(request.form.get('odometer'))
        start = float(request.form.get('start'))
        pumped = float(request.form.get('pumped'))
        end = start + pumped
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # ✅ Handle vehicle correctly for Plank vs Holfontein
        if site == "Plank":
            vehicle = request.form.get('vehicle_text')  # Free text input
        else:
            vehicle = request.form.get('vehicle_select')  # Dropdown

        # ✅ Validate start reading against previous end reading (1 decimal place)
        previous = get_last_readings()
        if previous:
            _, prev_end = previous
            if round(start, 1) != round(prev_end, 1):
                error = f"❌ Invalid entry: Start reading ({start:.1f}) must equal ({prev_end:.1f})."
                return render_template_string(HTML_FORM, error=error)

        # ✅ Only write to Excel if no error
        write_to_excel(timestamp, site, vehicle, driver_name, odometer, start, end, pumped)
        return render_template_string(HTML_FORM + "<p style='color:green; font-weight:bold;'>✅ Logged successfully!</p>")

    return render_template_string(HTML_FORM)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
